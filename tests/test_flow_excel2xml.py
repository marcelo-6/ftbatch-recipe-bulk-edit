# tests/test_flow_excel2xml.py

import os
import shutil
import pytest
from openpyxl import load_workbook
from lxml import etree

from core.parser import XMLParser
from core.exporter import ExcelExporter
from core.importer import ExcelImporter
from core.writer import XMLWriter
from core.base import NAMESPACE

SAMPLE_XML = f"""<?xml version="1.0" encoding="UTF-8"?>
<RecipeElement xmlns="{NAMESPACE}">
  <RecipeElementID>TEST</RecipeElementID>
  <Parameter>
    <Name>Param1</Name>
    <ERPAlias/>
    <PLCReference>1</PLCReference>
    <Real>0</Real>
    <High>100</High>
    <Low>0</Low>
    <EngineeringUnits/>
    <Scale>false</Scale>
  </Parameter>
  <Parameter>
    <Name>Param2</Name>
    <ERPAlias/>
    <PLCReference>1</PLCReference>
    <Integer>5</Integer>
    <High>10</High>
    <Low>1</Low>
    <EngineeringUnits/>
    <Scale>false</Scale>
  </Parameter>
  <Steps>
    <Step>
      <Name>Step1</Name>
      <FormulaValue>
        <Name>FV1</Name>
        <Display>false</Display>
        <Value/>
        <Integer>2</Integer>
        <EngineeringUnits/>
        <FormulaValueLimit Verification="No_Limits">
          <LowLowLowValue>0.</LowLowLowValue>
          <LowLowValue>0.</LowLowValue>
          <LowValue>0.</LowValue>
          <HighValue>0.</HighValue>
          <HighHighValue>0.</HighHighValue>
          <HighHighHighValue>0.</HighHighHighValue>
        </FormulaValueLimit>
      </FormulaValue>
    </Step>
  </Steps>
</RecipeElement>
"""


@pytest.fixture
def sample_pxml(tmp_path):
    p = tmp_path / "TEST.pxml"
    p.write_text(SAMPLE_XML, encoding="utf-8")
    return str(p)


def local_name(el):
    """Strip namespace and return just the tag's local name."""
    return el.tag.split("}", 1)[-1]


def test_excel2xml_full_flow(sample_pxml, tmp_path):
    """
    End-to-end excel2xml workflow test.

    This test exercises the full round-trip: exporting XML to Excel,
    programmatically modifying the spreadsheet, and re-importing back
    into updated XML files.

    It starts by exporting a sample `.pxml` to Excel, then alters the
    workbook by updating one parameter's value, deleting another row,
    and appending a new parameter with a valid data type.  After saving
    these changes, it invokes ExcelImporter to apply create/update/delete
    operations to the in-memory RecipeTree, capturing statistics of how
    many nodes were created, updated, or removed.

    Finally, XMLWriter serializes the modified trees into a timestamped
    output directory, and the test parses the resulting XML to assert that
    the update took effect (new value), the deletion occurred, and the new
    parameter exists.  This comprehensive test ensures the excel2xml path
    maintains XML integrity and fully honors user edits.
    """
    # 1) xml2excel
    parser = XMLParser()
    trees = parser.parse(sample_pxml)
    assert len(trees) == 1
    exporter = ExcelExporter()
    excel_file = tmp_path / "out.xlsx"
    exporter.export(trees, str(excel_file))

    # 2) modify Excel: update Param1.Real, delete Param2, add Param3 with a Real type
    wb = load_workbook(str(excel_file))
    ws = wb["TEST.pxml"]
    header = [c.value for c in ws[1]]
    idx_name = header.index("Name")
    idx_real = header.index("Real")

    # Update and delete existing rows
    for row in list(ws.iter_rows(min_row=2)):
        name = row[idx_name].value
        if name == "Param1":
            row[idx_real].value = "42"
        elif name == "Param2":
            ws.delete_rows(row[0].row)

    # Now append Param3 *with* a Real type (default 0)
    new_row = {col: "" for col in header}
    new_row["TagType"] = "Parameter"
    new_row["Name"] = "Param3"
    new_row["FullPath"] = "TEST/Parameter[Param3]"
    new_row["Real"] = "0"
    # Build the actual list in header order
    ws.append([new_row[col] for col in header])
    wb.save(str(excel_file))

    # 3) excel2xml
    importer = ExcelImporter()
    stats = importer.import_changes(str(excel_file), trees)
    assert stats["updated"] == 1
    assert stats["deleted"] == 1
    assert stats["created"] == 1

    # 4) write updated XMLs
    writer = XMLWriter()
    out_dir = writer.write(trees, base_dir=str(tmp_path))
    out_file = os.path.join(out_dir, "TEST.pxml")
    assert os.path.exists(out_file)

    # 5) verify resulting XML
    tree = etree.parse(out_file)
    root = tree.getroot()
    ns = {"ns": NAMESPACE}

    # Param1.Real == 42
    val1 = root.find("ns:Parameter[ns:Name='Param1']/ns:Real", namespaces=ns).text
    assert val1 == "42"

    # Param2 is removed
    assert root.find("ns:Parameter[ns:Name='Param2']", namespaces=ns) is None

    # Param3 exists
    p3 = root.find("ns:Parameter[ns:Name='Param3']", namespaces=ns)
    assert p3 is not None
    assert p3.find("ns:Name", namespaces=ns).text == "Param3"

    # ensure Param3 is inserted immediately after existing Parameters
    children = list(root)

    # helper:
    def local(el):
        return el.tag.split("}", 1)[-1]

    # find indexes
    param1_idx = next(
        i
        for i, el in enumerate(children)
        if local(el) == "Parameter"
        and el.find("ns:Name", namespaces=ns).text == "Param1"
    )
    param3_idx = next(
        i
        for i, el in enumerate(children)
        if local(el) == "Parameter"
        and el.find("ns:Name", namespaces=ns).text == "Param3"
    )
    steps_idx = next(i for i, el in enumerate(children) if local(el) == "Steps")

    # Assert ordering
    assert param1_idx < param3_idx < steps_idx

    # Cleanup
    shutil.rmtree(out_dir)
