import os
import shutil
import pytest
from openpyxl import load_workbook
from lxml import etree

from core.parser import XMLParser
from core.exporter import ExcelExporter
from core.importer import ExcelImporter
from core.writer import XMLWriter
from core.base import NAMESPACE

PARAM_EXPR_XML = f"""<?xml version="1.0" encoding="UTF-8"?>
<RecipeElement xmlns="{NAMESPACE}">
  <RecipeElementID>EXPR</RecipeElementID>
  <Steps>
    <Step>
      <Name>StepExpr</Name>
      <FormulaValue>
        <Name>FV_EXPR</Name>
        <Display>false</Display>
        <ParamExpression/>
        <Real>CP.GENE.EXPR * 2</Real>
        <EngineeringUnits>mL</EngineeringUnits>
      </FormulaValue>
    </Step>
  </Steps>
</RecipeElement>
"""


@pytest.fixture
def expr_pxml(tmp_path):
    p = tmp_path / "EXPR.pxml"
    p.write_text(PARAM_EXPR_XML, encoding="utf-8")
    return str(p)


def test_xml2excel_paramexpression(expr_pxml, tmp_path):
    """Export picks up ParamExpression and flags Real column appropriately."""
    trees = XMLParser().parse(expr_pxml)
    out = tmp_path / "expr.xlsx"
    ExcelExporter().export(trees, str(out))

    wb = load_workbook(str(out))
    ws = wb["EXPR.pxml"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # columns
    real_idx = header.index("Real")
    expr_idx = header.index("ParamExpression")

    # only one data row
    row = next(ws.iter_rows(min_row=2, values_only=True))
    # Real column must be literal "ParamExpression"
    # assert row[real_idx] == "ParamExpression"
    # ParamExpression column must carry the actual formula text
    assert row[real_idx] == "CP.GENE.EXPR * 2"


def test_excel2xml_paramexpression_roundtrip(expr_pxml, tmp_path):
    """Import recreates <ParamExpression/> and embeds formula into <Real>."""
    # 1) export
    trees = XMLParser().parse(expr_pxml)
    xlsx = tmp_path / "expr.xlsx"
    ExcelExporter().export(trees, str(xlsx))

    # 2) no manual edits needed—just import back
    stats = ExcelImporter().import_changes(str(xlsx), trees)
    # no create/delete, just update or skip
    assert stats["created"] == 0
    assert stats["deleted"] == 0
    # updated==0 since nothing changed
    assert stats["updated"] == 0

    # 3) write and verify XML
    out_dir = XMLWriter().write(trees, base_dir=str(tmp_path))
    result = etree.parse(os.path.join(out_dir, "EXPR.pxml"))
    root = result.getroot()
    ns = {"ns": NAMESPACE}

    # must have a <ParamExpression/> tag
    fv = root.find(".//ns:FormulaValue", namespaces=ns)
    assert fv.find("ns:ParamExpression", namespaces=ns) is not None

    # formula must live as text in the <Real> element
    real = fv.find("ns:Real", namespaces=ns).text
    assert real == "CP.GENE.EXPR * 2"

    shutil.rmtree(out_dir)
