#!/usr/bin/env python3

import os
import sys
import logging
import argparse
import datetime
from openpyxl import Workbook, load_workbook
from lxml import etree


###############################################################################
# GLOBAL ORDER DEFINITIONS (SCHEMA-BASED)
###############################################################################
# These lists define the order in which sub-elements should appear
# for new <Parameter> or <FormulaValue> nodes, as per your S88 schema.
# Adjust as needed for your real tags.

PARAMETER_TAG_ORDER = [
    "Name",
    "ERPAlias",
    "PLCReference",
    "Real",
    "High",
    "Low",
    "EngineeringUnits",
    "Scale",
]
FORMULAVALUE_TAG_ORDER = [
    "Name",           # typically first
    "Display",
    "Value",
    "String",
    "Defer",
    "Real",
    "EngineeringUnits",
    "FormulaValueLimit",  # We'll handle sub-elements in a sub-block
]
FORMULAVALUE_LIMIT_ORDER = [
    # We'll handle attribute "Verification" first (if present),
    # then these sub-elements in this exact order:
    "LowLowLowValue",
    "LowLowValue",
    "LowValue",
    "HighValue",
    "HighHighValue",
    "HighHighHighValue",
    # etc. if needed
]


###############################################################################
# LOGGING SETUP
###############################################################################
logger = logging.getLogger("BatchBulkEditor")


def configure_logging(debug_mode: bool) -> None:
    """
    If debug_mode is True => log to file at DEBUG level, plus console at INFO.
    Otherwise => only console at INFO, no file.
    """
    logger.handlers.clear()
    if debug_mode:
        logger.setLevel(logging.DEBUG)
        fh = logging.FileHandler("batch_bulk_editor.log", mode="a", encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")
        fh.setFormatter(fh_formatter)
        logger.addHandler(fh)

        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        ch_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
        ch.setFormatter(ch_formatter)
        logger.addHandler(ch)
    else:
        logger.setLevel(logging.INFO)
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        ch_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
        ch.setFormatter(ch_formatter)
        logger.addHandler(ch)


###############################################################################
# FILE DISCOVERY
###############################################################################
def find_child_xml_files(root_element, current_directory):
    child_files = []
    steps = root_element.findall(".//{urn:Rockwell/MasterRecipe}Step", namespaces=root_element.nsmap)
    for step in steps:
        sid = step.find("{urn:Rockwell/MasterRecipe}StepRecipeID", namespaces=root_element.nsmap)
        if sid is not None and sid.text:
            base = sid.text.strip()
            for ext in [".pxml", ".uxml", ".oxml"]:
                candidate = os.path.join(current_directory, base + ext)
                if os.path.isfile(candidate):
                    child_files.append(candidate)
                    break
    return child_files


###############################################################################
# SCHEMA-ORDER UTILS
###############################################################################
def insert_sub_element_in_order(parent, tagname, text, schema_list):
    """
    Insert <tagname> with text at the correct position among parent's children,
    according to the order in schema_list. If parent's children already exist,
    we place it after the last known schema-based sibling. If there's no known position,
    we place it at the end.
    """
    ns = parent.nsmap
    existing_children = list(parent)
    # figure out the index for insertion
    # we look for the schema_list index of 'tagname'
    if tagname not in schema_list:
        # not recognized => put at the end
        idx = len(existing_children)
    else:
        target_idx = schema_list.index(tagname)
        # we find the last child whose schema index is <= target_idx
        idx = 0
        last_eligible_pos = -1
        for i, child in enumerate(existing_children):
            child_local = etree.QName(child.tag).localname
            if child_local not in schema_list:
                continue
            child_idx = schema_list.index(child_local)
            if child_idx <= target_idx:
                last_eligible_pos = i
        idx = last_eligible_pos + 1

    new_el = etree.Element(f"{{urn:Rockwell/MasterRecipe}}{tagname}")
    new_el.text = text
    parent.insert(idx, new_el)
    return new_el


###############################################################################
# LOADING & RECURSING
###############################################################################
def load_and_recurse_xml(filepath, visited, all_params, stats):
    if filepath in visited:
        return
    visited.add(filepath)

    logger.debug(f"Parsing XML file: {filepath}")
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(filepath, parser)

    if filepath not in all_params:
        all_params[filepath] = []

    extract_parameters_from_root(tree, filepath, all_params, stats)

    root = tree.getroot()
    current_dir = os.path.dirname(filepath)
    child_xmls = find_child_xml_files(root, current_dir)
    for c in child_xmls:
        load_and_recurse_xml(c, visited, all_params, stats)


###############################################################################
# EXPORT (XML -> EXCEL)
###############################################################################
def build_full_parameter_path(recipe_id, parent_path_parts, tag_type, name_value):
    path = recipe_id
    for p in parent_path_parts:
        path += "/" + p
    path += f"/{tag_type}[{name_value}]"
    return path


def extract_parameters_from_root(tree, filepath, all_params, stats):
    root = tree.getroot()
    rec_id_el = root.find("{urn:Rockwell/MasterRecipe}RecipeElementID", namespaces=root.nsmap)
    recipe_id = rec_id_el.text.strip() if rec_id_el is not None else os.path.splitext(os.path.basename(filepath))[0]

    # top-level <Parameter>
    for param_el in root.findall("{urn:Rockwell/MasterRecipe}Parameter", namespaces=root.nsmap):
        row = extract_single_parameter(param_el, recipe_id, [])
        all_params[filepath].append(row)
        stats["parsed_count"] += 1
        logger.debug(f"Parsed Parameter: {row['FullPath']}")

    # steps => formula
    steps_el = root.find("{urn:Rockwell/MasterRecipe}Steps", namespaces=root.nsmap)
    if steps_el is not None:
        walk_step_structure(steps_el, recipe_id, ["Steps"], all_params[filepath], stats)


def walk_step_structure(steps_el, recipe_id, parent_path_parts, out_list, stats):
    for step_el in steps_el.findall("{urn:Rockwell/MasterRecipe}Step", namespaces=steps_el.nsmap):
        step_name_el = step_el.find("{urn:Rockwell/MasterRecipe}Name", namespaces=step_el.nsmap)
        step_name = step_name_el.text.strip() if (step_name_el is not None and step_name_el.text) else "UnknownStep"
        this_path = parent_path_parts + [f"Step[{step_name}]"]

        for fv_el in step_el.findall("{urn:Rockwell/MasterRecipe}FormulaValue", namespaces=step_el.nsmap):
            row = extract_single_formulavalue(fv_el, recipe_id, this_path)
            out_list.append(row)
            stats["parsed_count"] += 1
            logger.debug(f"Parsed FormulaValue: {row['FullPath']}")

        nested = step_el.find("{urn:Rockwell/MasterRecipe}Steps", namespaces=step_el.nsmap)
        if nested is not None:
            walk_step_structure(nested, recipe_id, this_path + ["Steps"], out_list, stats)


def extract_single_parameter(param_el, recipe_id, parent_path):
    row = {}
    row["TagType"] = "Parameter"
    name_el = param_el.find("{urn:Rockwell/MasterRecipe}Name", namespaces=param_el.nsmap)
    param_name = name_el.text.strip() if (name_el is not None and name_el.text) else ""
    row["Name"] = param_name
    row["FullPath"] = build_full_parameter_path(recipe_id, parent_path, "Parameter", param_name)

    for child in param_el:
        ln = etree.QName(child.tag).localname
        if ln == "Name":
            continue
        text_val = child.text if child.text is not None else ""
        row[ln] = text_val

    return row


def extract_single_formulavalue(fv_el, recipe_id, parent_path):
    row = {}
    row["TagType"] = "FormulaValue"
    name_el = fv_el.find("{urn:Rockwell/MasterRecipe}Name", namespaces=fv_el.nsmap)
    fv_name = name_el.text.strip() if (name_el is not None and name_el.text) else ""
    row["Name"] = fv_name
    row["FullPath"] = build_full_parameter_path(recipe_id, parent_path, "FormulaValue", fv_name)

    for child in fv_el:
        ln = etree.QName(child.tag).localname
        if ln == "Name":
            continue
        if ln == "FormulaValueLimit":
            ver = child.attrib.get("Verification", "")
            row["FormulaValueLimit_Verification"] = ver
            for sub in child:
                sub_ln = etree.QName(sub.tag).localname
                sub_text = sub.text if sub.text is not None else ""
                row[f"FormulaValueLimit_{sub_ln}"] = sub_text
        else:
            text_val = child.text if (child.text is not None) else ""
            row[ln] = text_val

    return row


def command_xml2excel(xml_path, excel_path, debug_mode=False):
    logger.info(f"Starting xml2excel: xml={xml_path}, excel={excel_path}, debug={debug_mode}")
    if not os.path.isfile(xml_path):
        logger.error(f"File not found: {xml_path}")
        sys.exit(1)

    visited = set()
    all_params = {}
    stats = {"parsed_count": 0}
    load_and_recurse_xml(xml_path, visited, all_params, stats)

    logger.info(f"Found total {stats['parsed_count']} parameters/formulas across {len(all_params)} file(s).")

    wb = Workbook()
    # remove default
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for filename, rows in all_params.items():
        short_name = os.path.splitext(os.path.basename(filename))[0]
        ws = wb.create_sheet(title=short_name[:31])

        # gather all columns
        all_cols = set()
        for r in rows:
            for k in r:
                all_cols.add(k)

        # convert to list, reorder for humans if you want, or define your own forced front columns
        # For simplicity, let's do something similar to before:
        all_cols_list = list(all_cols)

        # We'll just place FullPath, TagType, Name at the front
        forced_front = ["FullPath", "TagType", "Name"]
        ordered_cols = []
        for col in forced_front:
            if col in all_cols_list:
                ordered_cols.append(col)
                all_cols_list.remove(col)
        all_cols_list.sort()
        ordered_cols.extend(all_cols_list)

        ws.append(ordered_cols)
        for row_dict in rows:
            row_values = [row_dict.get(col, "") for col in ordered_cols]
            ws.append(row_values)

    wb.save(excel_path)
    logger.info("xml2excel complete. Workbook saved.")


###############################################################################
# IMPORT (EXCEL -> XML)
###############################################################################
def load_original_trees(xml_path):
    visited = set()
    file_to_tree = {}

    def _load_recursive(fp):
        if fp in visited:
            return
        visited.add(fp)
        parser = etree.XMLParser(remove_blank_text=False)
        t = etree.parse(fp, parser)
        file_to_tree[fp] = t
        root = t.getroot()
        child_paths = find_child_xml_files(root, os.path.dirname(fp))
        for c in child_paths:
            _load_recursive(c)

    _load_recursive(xml_path)
    return file_to_tree


def build_fullpath_map_for_file(tree, filepath, fullpath_map):
    """
    For each <Parameter> / <FormulaValue>, record node info. Also note sub-element text if needed for partial updates.
    """
    root = tree.getroot()
    rec_id_el = root.find("{urn:Rockwell/MasterRecipe}RecipeElementID", namespaces=root.nsmap)
    recipe_id = rec_id_el.text.strip() if rec_id_el is not None else os.path.splitext(os.path.basename(filepath))[0]

    # top-level <Parameter>
    for param_el in root.findall("{urn:Rockwell/MasterRecipe}Parameter", namespaces=root.nsmap):
        name_el = param_el.find("{urn:Rockwell/MasterRecipe}Name", namespaces=param_el.nsmap)
        param_name = name_el.text.strip() if (name_el is not None and name_el.text) else ""
        fp = build_full_parameter_path(recipe_id, [], "Parameter", param_name)

        ni = {
            "TagType": "Parameter",
            "Element": param_el,
            "OriginalSubs": {},
        }
        for child in param_el:
            ln = etree.QName(child.tag).localname
            if ln == "Name":
                continue
            ni["OriginalSubs"][ln] = child.text if child.text is not None else ""
        fullpath_map[fp] = ni

    # steps => formula
    steps_el = root.find("{urn:Rockwell/MasterRecipe}Steps", namespaces=root.nsmap)
    if steps_el is not None:
        walk_step_fullpath(steps_el, recipe_id, ["Steps"], fullpath_map)


def walk_step_fullpath(steps_el, recipe_id, parent_path_parts, fullpath_map):
    for step_el in steps_el.findall("{urn:Rockwell/MasterRecipe}Step", namespaces=steps_el.nsmap):
        name_el = step_el.find("{urn:Rockwell/MasterRecipe}Name", namespaces=step_el.nsmap)
        step_name = name_el.text.strip() if (name_el is not None and name_el.text) else "UnknownStep"
        this_path = parent_path_parts + [f"Step[{step_name}]"]

        for fv_el in step_el.findall("{urn:Rockwell/MasterRecipe}FormulaValue", namespaces=step_el.nsmap):
            fv_name_el = fv_el.find("{urn:Rockwell/MasterRecipe}Name", namespaces=fv_el.nsmap)
            fv_name = fv_name_el.text.strip() if (fv_name_el is not None and fv_name_el.text) else ""
            fp = build_full_parameter_path(recipe_id, this_path, "FormulaValue", fv_name)

            ni = {
                "TagType": "FormulaValue",
                "Element": fv_el,
                "OriginalSubs": {},  # for normal sub-elements
                "HasFormulaValueLimit": False,
                "FormulaValueLimitVerification": "",
                "FormulaValueLimitSubs": {},
            }
            for child in fv_el:
                ln = etree.QName(child.tag).localname
                if ln == "Name":
                    continue
                if ln == "FormulaValueLimit":
                    ni["HasFormulaValueLimit"] = True
                    ver = child.attrib.get("Verification", "")
                    ni["FormulaValueLimitVerification"] = ver
                    for sub in child:
                        sub_ln = etree.QName(sub.tag).localname
                        ni["FormulaValueLimitSubs"][sub_ln] = sub.text if sub.text is not None else ""
                else:
                    ni["OriginalSubs"][ln] = child.text if child.text is not None else ""
            fullpath_map[fp] = ni

        nested = step_el.find("{urn:Rockwell/MasterRecipe}Steps", namespaces=step_el.nsmap)
        if nested is not None:
            walk_step_fullpath(nested, recipe_id, this_path + ["Steps"], fullpath_map)


def remove_parameter_or_formulavalue(node_info, filepath, stats):
    el = node_info["Element"]
    p = el.getparent()
    if p is not None:
        p.remove(el)
        logger.info(f"Removed node from {filepath}: {etree.QName(el.tag).localname}")
        stats["deleted_count"] += 1


def locate_step_in_dom(root, middle_parts):
    current = root
    ns = root.nsmap
    for part in middle_parts:
        if part == "Steps":
            steps_el = current.find("{urn:Rockwell/MasterRecipe}Steps", namespaces=ns)
            if steps_el is None:
                return None
            current = steps_el
        elif part.startswith("Step["):
            sname = part[5:-1]
            found = None
            for s in current.findall("{urn:Rockwell/MasterRecipe}Step", namespaces=ns):
                nm = s.find("{urn:Rockwell/MasterRecipe}Name", namespaces=ns)
                if nm is not None and nm.text.strip() == sname:
                    found = s
                    break
            if not found:
                return None
            current = found
        else:
            return None
    return current


###############################################################################
# Creating / Updating with SCHEMA ORDER
###############################################################################
# def create_parameter_or_formulavalue(row_dict, filepath, tree, node_info, stats):
#     """
#     When we have a 'new' row in Excel, we create a brand-new <Parameter> or <FormulaValue>.
#     We will place sub-elements in the schema-defined order.
#     """
#     fp = row_dict.get("FullPath", "")
#     logger.info(f"Creating node in {filepath}: {fp}")

#     root = tree.getroot()
#     parts = fp.split("/")
#     last_part = parts[-1]
#     middle = parts[1:-1]

#     if last_part.startswith("Parameter["):
#         # top-level parameter => child of root
#         new_el = etree.Element("{urn:Rockwell/MasterRecipe}Parameter")
#         # We'll insert sub-elements in the known schema order
#         # 1) handle Name first
#         name_val = row_dict.get("Name", "").strip()
#         if name_val:
#             name_sub = etree.Element("{urn:Rockwell/MasterRecipe}Name")
#             name_sub.text = name_val
#             new_el.append(name_sub)

#         # 2) handle other known tags in order
#         for tag_name in PARAMETER_TAG_ORDER:
#             if tag_name in ("Name",):  # we did "Name" already
#                 continue
#             val = row_dict.get(tag_name, "").strip()
#             if val:
#                 insert_sub_element_in_order(new_el, tag_name, val, PARAMETER_TAG_ORDER)

#         # 3) handle any extra columns not in the official schema
#         for k, v in row_dict.items():
#             if k in ("FullPath", "TagType", "Name"):
#                 continue
#             if k in PARAMETER_TAG_ORDER:
#                 continue
#             vv = v.strip()
#             if vv:
#                 # We can just append or do a small insertion at the end
#                 insert_sub_element_in_order(new_el, k, vv, PARAMETER_TAG_ORDER)

#         # append to root
#         root.append(new_el)
#         stats["created_count"] += 1
#         logger.debug(f"Created Parameter => {fp}")

#     elif last_part.startswith("FormulaValue["):
#         # We must locate the <Step> node
#         step_el = locate_step_in_dom(root, middle)
#         if step_el is None:
#             logger.warning(f"Cannot locate step for {fp}, skipping create.")
#             return
#         new_el = etree.Element("{urn:Rockwell/MasterRecipe}FormulaValue")
#         # handle <Name>
#         name_val = row_dict.get("Name", "").strip()
#         if name_val:
#             name_sub = etree.Element("{urn:Rockwell/MasterRecipe}Name")
#             name_sub.text = name_val
#             new_el.append(name_sub)

#         # handle known tags in order, except FormulaValueLimit
#         for tag_name in FORMULAVALUE_TAG_ORDER:
#             if tag_name in ("Name", "FormulaValueLimit"):
#                 continue
#             val = row_dict.get(tag_name, "").strip()
#             if val:
#                 insert_sub_element_in_order(new_el, tag_name, val, FORMULAVALUE_TAG_ORDER)

#         # handle FormulaValueLimit if any columns are non-blank
#         fvl_keys = [k for k in row_dict if k.startswith("FormulaValueLimit_")]
#         if fvl_keys:
#             any_non_blank = any(row_dict[k].strip() != "" for k in fvl_keys)
#             if any_non_blank:
#                 fvl_el = etree.Element("{urn:Rockwell/MasterRecipe}FormulaValueLimit")
#                 # attribute
#                 ver = row_dict.get("FormulaValueLimit_Verification", "").strip()
#                 if ver:
#                     fvl_el.set("Verification", ver)

#                 # sub-elements in ORDER
#                 for subtag in FORMULAVALUE_LIMIT_ORDER:
#                     dict_key = f"FormulaValueLimit_{subtag}"
#                     sub_val = row_dict.get(dict_key, "").strip()
#                     if sub_val:
#                         sub_child = etree.Element(f"{{urn:Rockwell/MasterRecipe}}{subtag}")
#                         sub_child.text = sub_val
#                         fvl_el.append(sub_child)

#                 # place FVL in the right position
#                 insert_sub_element_in_order(new_el, "FormulaValueLimit", "", FORMULAVALUE_TAG_ORDER)
#                 # We need to replace the newly created empty node with our fvl_el
#                 # or we can do it differently. Let's do it carefully:
#                 # find the index
#                 for i, c in enumerate(new_el):
#                     if etree.QName(c.tag).localname == "FormulaValueLimit":
#                         new_el.remove(c)
#                         new_el.insert(i, fvl_el)
#                         break

#         # handle extra columns not recognized
#         known_keys = set(FORMULAVALUE_TAG_ORDER + fvl_keys)
#         for k, v in row_dict.items():
#             if k in known_keys or k in ("FullPath", "TagType", "Name"):
#                 continue
#             vv = v.strip()
#             if vv:
#                 # insert at the end
#                 insert_sub_element_in_order(new_el, k, vv, FORMULAVALUE_TAG_ORDER)

#         step_el.append(new_el)
#         stats["created_count"] += 1
#         logger.debug(f"Created FormulaValue => {fp}")

#     else:
#         logger.warning(f"Unrecognized path => {fp}")


def insert_sub_element_in_order(parent, tag_name, value, order_list):
    new_el = etree.Element(f"{{urn:Rockwell/MasterRecipe}}{tag_name}")
    new_el.text = value
    for i, child in enumerate(parent):
        if order_list and tag_name in order_list and order_list.index(tag_name) < order_list.index(etree.QName(child.tag).localname):
            parent.insert(i, new_el)
            return
    parent.append(new_el)

def locate_step_in_dom(root, path_parts):
    # Example function to locate a step element
    return root.find(".//Step")  # Placeholder

def create_parameter_or_formulavalue(row_dict, filepath, tree, node_info, stats):
    fp = row_dict.get("FullPath", "")
    logger.info(f"Creating node in {filepath}: {fp}")

    root = tree.getroot()
    parts = fp.split("/")
    last_part = parts[-1]
    middle = parts[1:-1]

    if last_part.startswith("Parameter["):
        new_el = etree.Element("{urn:Rockwell/MasterRecipe}Parameter")
        name_val = row_dict.get("Name", "").strip()
        if name_val:
            name_sub = etree.Element("{urn:Rockwell/MasterRecipe}Name")
            name_sub.text = name_val
            new_el.append(name_sub)
        
        for tag_name in PARAMETER_TAG_ORDER:
            if tag_name != "Name":
                val = row_dict.get(tag_name, "").strip()
                if val:
                    insert_sub_element_in_order(new_el, tag_name, val, PARAMETER_TAG_ORDER)
        
        for k, v in row_dict.items():
            if k not in PARAMETER_TAG_ORDER and k not in ("FullPath", "TagType", "Name"):
                vv = v.strip()
                if vv:
                    insert_sub_element_in_order(new_el, k, vv, PARAMETER_TAG_ORDER)
        
        insert_before = None
        for child in root:
            if etree.QName(child.tag).localname == "FormulaValue":
                insert_before = child
                break
        
        if insert_before is not None:
            root.insert(root.index(insert_before), new_el)
        else:
            root.append(new_el)
        
        stats["created_count"] += 1
        logger.debug(f"Created Parameter => {fp}")
    
    elif last_part.startswith("FormulaValue["):
        step_el = locate_step_in_dom(root, middle)
        if step_el is None:
            logger.warning(f"Cannot locate step for {fp}, skipping create.")
            return
        
        new_el = etree.Element("{urn:Rockwell/MasterRecipe}FormulaValue")
        name_val = row_dict.get("Name", "").strip()
        if name_val:
            name_sub = etree.Element("{urn:Rockwell/MasterRecipe}Name")
            name_sub.text = name_val
            new_el.append(name_sub)
        
        for tag_name in FORMULAVALUE_TAG_ORDER:
            if tag_name not in ("Name", "FormulaValueLimit"):
                val = row_dict.get(tag_name, "").strip()
                if val:
                    insert_sub_element_in_order(new_el, tag_name, val, FORMULAVALUE_TAG_ORDER)
        
        fvl_keys = [k for k in row_dict if k.startswith("FormulaValueLimit_")]
        if any(row_dict[k].strip() for k in fvl_keys):
            fvl_el = etree.Element("{urn:Rockwell/MasterRecipe}FormulaValueLimit")
            ver = row_dict.get("FormulaValueLimit_Verification", "").strip()
            if ver:
                fvl_el.set("Verification", ver)
            
            for subtag in FORMULAVALUE_LIMIT_ORDER:
                sub_val = row_dict.get(f"FormulaValueLimit_{subtag}", "").strip()
                if sub_val:
                    sub_child = etree.Element(f"{{urn:Rockwell/MasterRecipe}}{subtag}")
                    sub_child.text = sub_val
                    fvl_el.append(sub_child)
            
            insert_sub_element_in_order(new_el, "FormulaValueLimit", "", FORMULAVALUE_TAG_ORDER)
            for i, c in enumerate(new_el):
                if etree.QName(c.tag).localname == "FormulaValueLimit":
                    new_el.remove(c)
                    new_el.insert(i, fvl_el)
                    break
        
        known_keys = set(FORMULAVALUE_TAG_ORDER + fvl_keys)
        for k, v in row_dict.items():
            if k not in known_keys and k not in ("FullPath", "TagType", "Name"):
                vv = v.strip()
                if vv:
                    insert_sub_element_in_order(new_el, k, vv, FORMULAVALUE_TAG_ORDER)
        
        step_el.append(new_el)
        stats["created_count"] += 1
        logger.debug(f"Created FormulaValue => {fp}")
    else:
        logger.warning(f"Unrecognized path => {fp}")
        

def update_parameter_or_formulavalue(node_info, row_dict, filepath, stats):
    """
    For existing nodes, partially update sub-elements while preserving:
    1) The node's existing child order
    2) The node's existing sub-element text if user left it blank
    3) Insert new sub-elements in the correct schema order if user has a new value
    """
    el = node_info["Element"]
    tagtype = node_info["TagType"]
    if tagtype == "Parameter":
        partial_update_parameter(el, row_dict, node_info)
        logger.debug(f"Updated Parameter: {row_dict.get('FullPath','')}")
    else:
        partial_update_formulavalue(el, row_dict, node_info)
        logger.debug(f"Updated FormulaValue: {row_dict.get('FullPath','')}")
    stats["updated_count"] += 1


def partial_update_parameter(param_element, row_dict, node_info):
    ns = param_element.nsmap
    original_subs = node_info["OriginalSubs"]  # localName-> old text

    # We'll define a helper that tries to find existing child or insert in order
    def set_or_insert_parameter_sub(parent, tag_name, val):
        # find if it already exists
        child = parent.find(f"{{urn:Rockwell/MasterRecipe}}{tag_name}", namespaces=ns)
        if child is not None:
            child.text = val
        else:
            # insert in correct order
            insert_sub_element_in_order(parent, tag_name, val, PARAMETER_TAG_ORDER)

    # For each known tag in schema order, see what user typed
    # but do partial logic
    for tag_name in PARAMETER_TAG_ORDER:
        if tag_name == "Name":
            continue  # We do not remove or reorder <Name>, it's already there
        new_val = row_dict.get(tag_name, "")
        new_val_strip = new_val.strip()
        old_val = original_subs.get(tag_name, None)
        if new_val_strip == "":
            # user left blank
            if old_val is None:
                # didn't exist => skip
                pass
            else:
                # existed => preserve old
                pass
        else:
            # user gave something => set or create
            set_or_insert_parameter_sub(param_element, tag_name, new_val_strip)

    # handle any columns not in schema
    for k, v in row_dict.items():
        if k in ("FullPath", "TagType", "Name"):
            continue
        if k in PARAMETER_TAG_ORDER:
            continue
        new_val_strip = v.strip()
        old_val = original_subs.get(k, None)
        if new_val_strip == "":
            if old_val is None:
                pass
            else:
                # preserve
                pass
        else:
            # insert or set
            set_or_insert_parameter_sub(param_element, k, new_val_strip)


def partial_update_formulavalue(fv_element, row_dict, node_info):
    ns = fv_element.nsmap
    original_subs = node_info["OriginalSubs"]  # normal sub-elements
    had_fvl = node_info["HasFormulaValueLimit"]
    old_ver = node_info["FormulaValueLimitVerification"]
    old_subs = node_info["FormulaValueLimitSubs"]

    def set_or_insert_formulavalue_sub(parent, tag_name, val):
        existing_child = parent.find(f"{{urn:Rockwell/MasterRecipe}}{tag_name}", namespaces=ns)
        if existing_child is not None:
            existing_child.text = val
        else:
            insert_sub_element_in_order(parent, tag_name, val, FORMULAVALUE_TAG_ORDER)

    # 1) handle normal sub-elements (non-FVL)
    for tag_name in FORMULAVALUE_TAG_ORDER:
        if tag_name in ("Name", "FormulaValueLimit"):
            continue
        new_val = row_dict.get(tag_name, "")
        new_val_strip = new_val.strip()
        old_val = original_subs.get(tag_name, None)
        if new_val_strip == "":
            if old_val is None:
                pass
            else:
                pass
        else:
            set_or_insert_formulavalue_sub(fv_element, tag_name, new_val_strip)

    # also handle any weird columns not recognized
    recognized_keys = set(FORMULAVALUE_TAG_ORDER)
    fvl_keys = [k for k in row_dict if k.startswith("FormulaValueLimit_")]
    recognized_keys.update(fvl_keys)
    for k, v in row_dict.items():
        if k in ("FullPath", "TagType", "Name"):
            continue
        if k in recognized_keys:
            continue
        vv = v.strip()
        old_val = original_subs.get(k, None)
        if vv == "":
            if old_val is None:
                pass
            else:
                pass
        else:
            set_or_insert_formulavalue_sub(fv_element, k, vv)

    # 2) handle <FormulaValueLimit> partial
    if not fvl_keys:
        # user didn't mention FVL => do nothing
        return

    fvl_el = fv_element.find("{urn:Rockwell/MasterRecipe}FormulaValueLimit", namespaces=ns)
    any_non_blank = any(row_dict[k].strip() != "" for k in fvl_keys)
    if not had_fvl and fvl_el is None:
        if any_non_blank:
            # create FVL
            # but place it in correct spot
            insert_sub_element_in_order(fv_element, "FormulaValueLimit", "", FORMULAVALUE_TAG_ORDER)
            # now we find that node
            fvl_el = fv_element.find("{urn:Rockwell/MasterRecipe}FormulaValueLimit", namespaces=ns)
    elif had_fvl and fvl_el is None:
        # recreate from old data if we want to preserve?
        fvl_el = etree.Element("{urn:Rockwell/MasterRecipe}FormulaValueLimit")
        if old_ver:
            fvl_el.set("Verification", old_ver)
        for subk, subv in old_subs.items():
            sub_child = etree.Element(f"{{urn:Rockwell/MasterRecipe}}{subk}")
            sub_child.text = subv
            fvl_el.append(sub_child)
        insert_sub_element_in_order(fv_element, "FormulaValueLimit", "", FORMULAVALUE_TAG_ORDER)
        new_fvl = fv_element.find("{urn:Rockwell/MasterRecipe}FormulaValueLimit", namespaces=ns)
        fv_element.replace(new_fvl, fvl_el)

    if fvl_el is None:
        return

    # update partial
    new_ver = row_dict.get("FormulaValueLimit_Verification", "").strip()
    if new_ver == "":
        if had_fvl:
            # keep old
            pass
    else:
        fvl_el.set("Verification", new_ver)

    for subtag in FORMULAVALUE_LIMIT_ORDER:
        dict_key = f"FormulaValueLimit_{subtag}"
        new_subval = row_dict.get(dict_key, "").strip()
        old_v = old_subs.get(subtag, None) if had_fvl else None
        if new_subval == "":
            if old_v is None:
                pass
            else:
                pass
        else:
            # create or set
            existing_sub = fvl_el.find(f"{{urn:Rockwell/MasterRecipe}}{subtag}", namespaces=ns)
            if existing_sub is None:
                # place it in correct order
                # We'll do a simple approach: append in order
                # We'll find the current sub children, see where to place it
                idx = 0
                existing_fvl_children = list(fvl_el)
                # figure out subtag's index
                target_idx = FORMULAVALUE_LIMIT_ORDER.index(subtag)
                last_ok_pos = -1
                for i, c in enumerate(existing_fvl_children):
                    child_local = etree.QName(c.tag).localname
                    if child_local in FORMULAVALUE_LIMIT_ORDER:
                        c_idx = FORMULAVALUE_LIMIT_ORDER.index(child_local)
                        if c_idx <= target_idx:
                            last_ok_pos = i
                idx = last_ok_pos + 1
                new_child = etree.Element(f"{{urn:Rockwell/MasterRecipe}}{subtag}")
                new_child.text = new_subval
                fvl_el.insert(idx, new_child)
            else:
                existing_sub.text = new_subval


def command_excel2xml(xml_path, excel_path, debug_mode=False):
    logger.info(f"Starting excel2xml: xml={xml_path}, excel={excel_path}, debug={debug_mode}")
    if not os.path.isfile(xml_path):
        logger.error(f"File not found: {xml_path}")
        sys.exit(1)
    if not os.path.isfile(excel_path):
        logger.error(f"Excel file not found: {excel_path}")
        sys.exit(1)

    file_to_tree = load_original_trees(xml_path)
    file_to_fullpath_map = {}
    for fpath, tree in file_to_tree.items():
        fm = {}
        build_fullpath_map_for_file(tree, fpath, fm)
        file_to_fullpath_map[fpath] = fm

    wb = load_workbook(excel_path)
    sheet_to_data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = rows[0]
        data_rows = rows[1:]
        this_sheet_data = {}
        for dr in data_rows:
            row_dict = {}
            for col_idx, col_name in enumerate(headers):
                if not col_name:
                    continue
                val = dr[col_idx] if col_idx < len(dr) else ""
                if val is None:
                    val = ""
                row_dict[col_name] = str(val)
            fp = row_dict.get("FullPath", "")
            if fp:
                this_sheet_data[fp] = row_dict
        sheet_to_data[sheet] = this_sheet_data

    stats = {
        "deleted_count": 0,
        "created_count": 0,
        "updated_count": 0,
    }

    for fpath, tree in file_to_tree.items():
        short_name = os.path.splitext(os.path.basename(fpath))[0]
        excel_data_for_file = sheet_to_data.get(short_name, {})
        existing_map = file_to_fullpath_map[fpath]

        excel_fullpaths = set(excel_data_for_file.keys())
        existing_fullpaths = set(existing_map.keys())

        to_delete = existing_fullpaths - excel_fullpaths
        to_create = excel_fullpaths - existing_fullpaths
        to_update = existing_fullpaths.intersection(excel_fullpaths)

        for fp in to_delete:
            remove_parameter_or_formulavalue(existing_map[fp], fpath, stats)
        for fp in to_create:
            row_d = excel_data_for_file[fp]
            create_parameter_or_formulavalue(row_d, fpath, tree, node_info=None, stats=stats)
        for fp in to_update:
            ni = existing_map[fp]
            row_d = excel_data_for_file[fp]
            update_parameter_or_formulavalue(ni, row_d, fpath, stats)

    logger.info(f"Excel2XML summary => Deleted: {stats['deleted_count']}, Created: {stats['created_count']}, Updated: {stats['updated_count']}")

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d-%H%M%S")
    outdir = os.path.join(os.path.dirname(xml_path), "converted-outputs", timestamp)
    os.makedirs(outdir, exist_ok=True)

    for fpath, tree in file_to_tree.items():
        outpath = os.path.join(outdir, os.path.basename(fpath))
        logger.info(f"Writing updated XML => {outpath}")
        tree.write(outpath, encoding="UTF-8", xml_declaration=True, pretty_print=False)

    logger.info("excel2xml complete. All updated XML saved.")


###############################################################################
# MAIN
###############################################################################
def main():
    parser = argparse.ArgumentParser(description="FactoryTalk Batch Bulk Editor")
    parser.add_argument("--debug", action="store_true", default=False,
                        help="Enable debug logging to file (batch_bulk_editor.log).")

    sub = parser.add_subparsers(dest="command", help="Sub-command")

    p1 = sub.add_parser("xml2excel", help="Export .pxml/.uxml/.oxml + children to Excel")
    p1.add_argument("--xml", required=True, help="Parent .pxml/.uxml/.oxml file")
    p1.add_argument("--excel", required=True, help="Output .xlsx path")

    p2 = sub.add_parser("excel2xml", help="Import Excel changes back to .pxml/.uxml/.oxml")
    p2.add_argument("--xml", required=True, help="Parent .pxml/.uxml/.oxml")
    p2.add_argument("--excel", required=True, help="Edited .xlsx path")

    args = parser.parse_args()
    configure_logging(args.debug)

    if args.command == "xml2excel":
        command_xml2excel(args.xml, args.excel, debug_mode=args.debug)
    elif args.command == "excel2xml":
        command_excel2xml(args.xml, args.excel, debug_mode=args.debug)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
