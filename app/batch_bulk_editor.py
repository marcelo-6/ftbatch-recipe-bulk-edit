#!/usr/bin/env python3

import os
import sys
import logging
import argparse
import datetime
from openpyxl import Workbook, load_workbook
from lxml import etree


# ----------------------------------------------------------------------
# GLOBAL COLUMN ORDER (for Excel output readability)
# ----------------------------------------------------------------------

DEFAULT_COLUMN_ORDER = [
    "FullPath",
    "TagType",
    "Name",
    "ERPAlias",
    "Display"
    "PLCReference",
    "Real",
    "Integer",  # If you use that in your real schema
    "String",
    "Defer",
    "High",
    "Low",
    "EnumerationSet",
    "EnumerationMember",
    "EngineeringUnits",
    "Scale",
    "FormulaValueLimit_Verification",
    "FormulaValueLimit_LowLowLowValue",
    "FormulaValueLimit_LowLowValue",
    "FormulaValueLimit_LowValue",
    "FormulaValueLimit_HighValue",
    "FormulaValueLimit_HighHighValue",
    "FormulaValueLimit_HighHighHighValue",
    # etc. Add or remove as your real use-case demands
]

# ----------------------------------------------------------------------
# LOGGING SETUP
# ----------------------------------------------------------------------

logger = logging.getLogger("BatchBulkEditor")


def configure_logging(debug_mode: bool) -> None:
    """
    If debug_mode is True:
      - root logger level = DEBUG
      - create file handler at DEBUG
      - console at INFO
    Otherwise:
      - root logger level = INFO
      - no file is created
      - console at INFO
    """
    # Remove any existing handlers if re-run
    logger.handlers.clear()

    if debug_mode:
        logger.setLevel(logging.DEBUG)
        # File handler at DEBUG
        fh = logging.FileHandler("batch_bulk_editor.log", mode="a", encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter(
            fmt="[%(asctime)s] %(levelname)s [%(name)s.%(funcName)s:%(lineno)d] %(message)s",
            datefmt="%d-%b-%Y %H:%M:%S",
        )
        fh.setFormatter(file_formatter)
        logger.addHandler(fh)

        # Console at INFO
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        console_formatter = logging.Formatter(
            fmt="[%(asctime)s] %(levelname)s: %(message)s",
            datefmt="%d-%b-%Y %H:%M:%S",
        )
        ch.setFormatter(console_formatter)
        logger.addHandler(ch)
    else:
        # No file handler, just console
        logger.setLevel(logging.INFO)
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        console_formatter = logging.Formatter(
            fmt="[%(asctime)s] %(levelname)s: %(message)s",
            datefmt="%d-%b-%Y %H:%M:%S",
        )
        ch.setFormatter(console_formatter)
        logger.addHandler(ch)


# ----------------------------------------------------------------------
# FIND CHILD FILES
# ----------------------------------------------------------------------


def find_child_xml_files(root_element, current_directory):
    """
    For <StepRecipeID> in <Step>, see if there's a .pxml, .uxml, or .oxml in same directory with that base name.
    """
    child_files = []
    steps = root_element.findall(
        ".//{urn:Rockwell/MasterRecipe}Step", namespaces=root_element.nsmap
    )

    for step in steps:
        sid = step.find(
            "{urn:Rockwell/MasterRecipe}StepRecipeID", namespaces=root_element.nsmap
        )
        if sid is not None and sid.text:
            base = sid.text.strip()
            for ext in [".PXML", ".UXML", ".OXML"]:
                candidate = os.path.join(current_directory, base + ext)
                if os.path.isfile(candidate):
                    child_files.append(candidate)
                    break
    return child_files


# ----------------------------------------------------------------------
# LOADING FILES (RECURSIVE)
# ----------------------------------------------------------------------


def load_and_recurse_xml(filepath, visited, all_params, stats):
    """
    Recursively parse parent + child files; store param info in all_params[filepath].
    Also fill in stats for logging (like how many we parse).
    """
    if filepath in visited:
        return
    visited.add(filepath)

    logger.debug(f"Parsing XML file: {filepath}")
    parser = etree.XMLParser(remove_blank_text=False)
    tree = etree.parse(filepath, parser)

    if filepath not in all_params:
        all_params[filepath] = []

    extract_parameters_from_root(tree, filepath, all_params, stats)

    root = tree.getroot()
    current_dir = os.path.dirname(filepath)
    child_xmls = find_child_xml_files(root, current_dir)

    for c in child_xmls:
        load_and_recurse_xml(c, visited, all_params, stats)


# ----------------------------------------------------------------------
# EXTRACTING (XML -> PYTHON ROWS)
# ----------------------------------------------------------------------


def build_full_parameter_path(recipe_id, parent_path_parts, tag_type, name_value):
    """
    e.g. "OP_LNP_SETUP/Parameter[AQREL_DEVICE_SETUP]"
         "OP_LNP_SETUP/Steps/Step[ACQ_REL:1]/FormulaValue[X_R_TARE_ZERO_TIME_OUT_SEC]"
    """
    path = recipe_id
    for p in parent_path_parts:
        path += "/" + p
    path += f"/{tag_type}[{name_value}]"
    return path


def extract_parameters_from_root(tree, filepath, all_params, stats):
    """
    Parse top-level <Parameter> and <FormulaValue> within <Steps>, store in all_params.
    Update stats for logging.
    """
    root = tree.getroot()
    rec_id_el = root.find(
        "{urn:Rockwell/MasterRecipe}RecipeElementID", namespaces=root.nsmap
    )
    recipe_id = (
        rec_id_el.text.strip()
        if rec_id_el is not None
        else os.path.splitext(os.path.basename(filepath))[0]
    )

    # top-level <Parameter>
    for param_el in root.findall(
        "{urn:Rockwell/MasterRecipe}Parameter", namespaces=root.nsmap
    ):
        row = extract_single_parameter(param_el, recipe_id, [])
        all_params[filepath].append(row)
        stats["parsed_count"] += 1
        logger.debug(f"Parsed Parameter: {row['FullPath']}")

    # steps => formula
    steps_el = root.find("{urn:Rockwell/MasterRecipe}Steps", namespaces=root.nsmap)
    if steps_el is not None:
        walk_step_structure(steps_el, recipe_id, ["Steps"], all_params[filepath], stats)


def walk_step_structure(steps_el, recipe_id, parent_path_parts, out_list, stats):
    for step_el in steps_el.findall(
        "{urn:Rockwell/MasterRecipe}Step", namespaces=steps_el.nsmap
    ):
        step_name_el = step_el.find(
            "{urn:Rockwell/MasterRecipe}Name", namespaces=step_el.nsmap
        )
        step_name = (
            step_name_el.text.strip()
            if (step_name_el is not None and step_name_el.text)
            else "UnknownStep"
        )
        this_path = parent_path_parts + [f"Step[{step_name}]"]

        for fv_el in step_el.findall(
            "{urn:Rockwell/MasterRecipe}FormulaValue", namespaces=step_el.nsmap
        ):
            row = extract_single_formulavalue(fv_el, recipe_id, this_path)
            out_list.append(row)
            stats["parsed_count"] += 1
            logger.debug(f"Parsed FormulaValue: {row['FullPath']}")

        # nested steps?
        nested = step_el.find(
            "{urn:Rockwell/MasterRecipe}Steps", namespaces=step_el.nsmap
        )
        if nested is not None:
            walk_step_structure(
                nested, recipe_id, this_path + ["Steps"], out_list, stats
            )


def extract_single_parameter(param_el, recipe_id, parent_path):
    row = {}
    row["TagType"] = "Parameter"
    name_el = param_el.find(
        "{urn:Rockwell/MasterRecipe}Name", namespaces=param_el.nsmap
    )
    param_name = name_el.text.strip() if (name_el is not None and name_el.text) else ""
    row["Name"] = param_name
    row["FullPath"] = build_full_parameter_path(
        recipe_id, parent_path, "Parameter", param_name
    )

    for child in param_el:
        if etree.QName(child.tag).localname == "Name":
            continue
        text_val = child.text if child.text is not None else ""
        row[etree.QName(child.tag).localname] = text_val

    return row


def extract_single_formulavalue(fv_el, recipe_id, parent_path):
    row = {}
    row["TagType"] = "FormulaValue"
    name_el = fv_el.find("{urn:Rockwell/MasterRecipe}Name", namespaces=fv_el.nsmap)
    fv_name = name_el.text.strip() if (name_el is not None and name_el.text) else ""
    row["Name"] = fv_name
    row["FullPath"] = build_full_parameter_path(
        recipe_id, parent_path, "FormulaValue", fv_name
    )

    for child in fv_el:
        ln = etree.QName(child.tag).localname
        if ln == "Name":
            continue
        if ln == "FormulaValueLimit":
            # store attribute "Verification" as "FormulaValueLimit_Verification"
            verification_val = child.attrib.get("Verification", "")
            row["FormulaValueLimit_Verification"] = verification_val

            for sub in child:
                sub_ln = etree.QName(sub.tag).localname
                sub_text = sub.text if sub.text is not None else ""
                row[f"FormulaValueLimit_{sub_ln}"] = sub_text
        else:
            text_val = child.text if (child.text is not None) else ""
            row[ln] = text_val

    return row


# ----------------------------------------------------------------------
# XML -> EXCEL
# ----------------------------------------------------------------------


# def command_xml2excel(parent_xml, excel_path):
#     logger.info(f"Starting xml2excel: parent={parent_xml} -> {excel_path}")
#     if not os.path.isfile(parent_xml):
#         logger.error(f"File not found: {parent_xml}")
#         sys.exit(1)

#     visited = set()
#     all_params = {}  # file->list of row-dicts
#     stats = {"parsed_count": 0}
#     load_and_recurse_xml(parent_xml, visited, all_params, stats)

#     logger.info(
#         f"Found total {stats['parsed_count']} parameters/formulas across {len(all_params)} file(s)."
#     )

#     wb = Workbook()
#     # remove default
#     if "Sheet" in wb.sheetnames:
#         wb.remove(wb["Sheet"])

#     for filename, rows in all_params.items():
#         short_name = os.path.splitext(os.path.basename(filename))[0]
#         ws = wb.create_sheet(title=short_name[:31])

#         # Gather all columns
#         all_cols = set()
#         for r in rows:
#             for k in r:
#                 all_cols.add(k)

#         # Convert to a list, reorder using DEFAULT_COLUMN_ORDER
#         all_cols_list = list(all_cols)
#         ordered_cols = []
#         # (1) Add in default columns in the specified order
#         for col in DEFAULT_COLUMN_ORDER:
#             if col in all_cols_list:
#                 ordered_cols.append(col)
#                 all_cols_list.remove(col)
#         # (2) Append leftover columns in alphabetical order
#         all_cols_list.sort()
#         ordered_cols.extend(all_cols_list)

#         ws.append(ordered_cols)
#         for row_dict in rows:
#             row_values = [row_dict.get(col, "") for col in ordered_cols]
#             ws.append(row_values)

#     wb.save(excel_path)
#     logger.info("xml2excel complete. Workbook saved.")


def command_xml2excel(parent_xml, excel_path):
    logger.info(f"Starting xml2excel: parent={parent_xml} -> {excel_path}")
    if not os.path.isfile(parent_xml):
        logger.error(f"File not found: {parent_xml}")
        sys.exit(1)

    visited = set()
    all_params = {}  # file->list of row-dicts
    stats = {"parsed_count": 0}
    load_and_recurse_xml(parent_xml, visited, all_params, stats)

    logger.info(
        f"Found total {stats['parsed_count']} parameters/formulas across {len(all_params)} file(s)."
    )

    wb = Workbook()
    # remove default
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for filename, rows in all_params.items():
        short_name = os.path.splitext(os.path.basename(filename))[0]
        ws = wb.create_sheet(title=short_name[:31])

        if not rows:
            continue
        
        # Maintain XML column order dynamically
        ordered_cols = []
        seen_cols = set()
        
        for row in rows:
            for col in row.keys():
                if col not in seen_cols:
                    ordered_cols.append(col)
                    seen_cols.add(col)

        ws.append(ordered_cols)
        for row_dict in rows:
            row_values = [row_dict.get(col, "") for col in ordered_cols]
            ws.append(row_values)

    wb.save(excel_path)
    logger.info("xml2excel complete. Workbook saved.")




# ----------------------------------------------------------------------
# EXCEL -> XML
# ----------------------------------------------------------------------


def load_original_trees(parent_xml):
    visited = set()
    file_to_tree = {}

    def _load_recursive(fp):
        if fp in visited:
            return
        visited.add(fp)

        parser = etree.XMLParser(remove_blank_text=False)
        t = etree.parse(fp, parser)
        file_to_tree[fp] = t

        root = t.getroot()
        child_paths = find_child_xml_files(root, os.path.dirname(fp))
        for c in child_paths:
            _load_recursive(c)

    _load_recursive(parent_xml)
    return file_to_tree


def build_fullpath_map_for_file(tree, filepath, fullpath_map):
    """
    For each <Parameter> or <FormulaValue>, record:
      - node_info["TagType"]
      - node_info["Element"] = the actual node
      - node_info["OriginalSubs"] = set/dict of sub-element localnames => original text
      - node_info["HasFormulaValueLimit"] = True if it originally had one
      - node_info["FormulaValueLimitSubs"] = dict of sub-element name => text
      - node_info["FormulaValueLimitVerification"] = original attribute
    """
    root = tree.getroot()
    rec_id_el = root.find(
        "{urn:Rockwell/MasterRecipe}RecipeElementID", namespaces=root.nsmap
    )
    recipe_id = (
        rec_id_el.text.strip()
        if rec_id_el is not None
        else os.path.splitext(os.path.basename(filepath))[0]
    )

    # top-level <Parameter>
    for param_el in root.findall(
        "{urn:Rockwell/MasterRecipe}Parameter", namespaces=root.nsmap
    ):
        name_el = param_el.find(
            "{urn:Rockwell/MasterRecipe}Name", namespaces=param_el.nsmap
        )
        param_name = (
            name_el.text.strip() if (name_el is not None and name_el.text) else ""
        )
        fp = build_full_parameter_path(recipe_id, [], "Parameter", param_name)

        ni = {
            "TagType": "Parameter",
            "Element": param_el,
            "OriginalSubs": {},
        }
        for child in param_el:
            ln = etree.QName(child.tag).localname
            if ln == "Name":
                continue
            ni["OriginalSubs"][ln] = child.text if child.text is not None else ""

        fullpath_map[fp] = ni

    # Steps => formula
    steps_el = root.find("{urn:Rockwell/MasterRecipe}Steps", namespaces=root.nsmap)
    if steps_el is not None:
        walk_step_fullpath(steps_el, recipe_id, ["Steps"], fullpath_map)


def walk_step_fullpath(steps_el, recipe_id, parent_path_parts, fullpath_map):
    for step_el in steps_el.findall(
        "{urn:Rockwell/MasterRecipe}Step", namespaces=steps_el.nsmap
    ):
        name_el = step_el.find(
            "{urn:Rockwell/MasterRecipe}Name", namespaces=step_el.nsmap
        )
        step_name = (
            name_el.text.strip()
            if (name_el is not None and name_el.text)
            else "UnknownStep"
        )
        this_path = parent_path_parts + [f"Step[{step_name}]"]

        for fv_el in step_el.findall(
            "{urn:Rockwell/MasterRecipe}FormulaValue", namespaces=step_el.nsmap
        ):
            fv_name_el = fv_el.find(
                "{urn:Rockwell/MasterRecipe}Name", namespaces=fv_el.nsmap
            )
            fv_name = (
                fv_name_el.text.strip()
                if (fv_name_el is not None and fv_name_el.text)
                else ""
            )
            fp = build_full_parameter_path(
                recipe_id, this_path, "FormulaValue", fv_name
            )

            ni = {
                "TagType": "FormulaValue",
                "Element": fv_el,
                "OriginalSubs": {},
                "HasFormulaValueLimit": False,
                "FormulaValueLimitVerification": "",
                "FormulaValueLimitSubs": {},
            }

            for child in fv_el:
                ln = etree.QName(child.tag).localname
                if ln == "Name":
                    continue
                if ln == "FormulaValueLimit":
                    ni["HasFormulaValueLimit"] = True
                    ver = child.attrib.get("Verification", "")
                    ni["FormulaValueLimitVerification"] = ver
                    for sub in child:
                        sub_ln = etree.QName(sub.tag).localname
                        ni["FormulaValueLimitSubs"][sub_ln] = (
                            sub.text if sub.text is not None else ""
                        )
                else:
                    ni["OriginalSubs"][ln] = (
                        child.text if child.text is not None else ""
                    )

            fullpath_map[fp] = ni

        nested = step_el.find(
            "{urn:Rockwell/MasterRecipe}Steps", namespaces=step_el.nsmap
        )
        if nested is not None:
            walk_step_fullpath(nested, recipe_id, this_path + ["Steps"], fullpath_map)


def remove_parameter_or_formulavalue(node_info, filepath, stats):
    el = node_info["Element"]
    p = el.getparent()
    if p is not None:
        p.remove(el)
        logger.info(f"Removed node from {filepath}: {etree.QName(el.tag).localname}")
        stats["deleted_count"] += 1


def locate_step_in_dom(root, middle_parts):
    current = root
    ns = root.nsmap
    for part in middle_parts:
        if part == "Steps":
            steps_el = current.find("{urn:Rockwell/MasterRecipe}Steps", namespaces=ns)
            if steps_el is None:
                return None
            current = steps_el
        elif part.startswith("Step["):
            sname = part[5:-1]
            found = None
            for s in current.findall("{urn:Rockwell/MasterRecipe}Step", namespaces=ns):
                nm = s.find("{urn:Rockwell/MasterRecipe}Name", namespaces=ns)
                if nm is not None and nm.text.strip() == sname:
                    found = s
                    break
            if not found:
                return None
            current = found
        else:
            return None
    return current


# def create_parameter_or_formulavalue(row_dict, filepath, tree, node_info, stats):
#     fp = row_dict.get("FullPath", "")
#     logger.info(f"Creating node in {filepath}: {fp}")

#     root = tree.getroot()
#     parts = fp.split("/")
#     last_part = parts[-1]
#     middle = parts[1:-1]

#     # Lists to store new elements
#     new_parameters = []
#     new_formula_values = []

#     if last_part.startswith("Parameter["):
#         # Create new parameter element
#         new_el = etree.Element("{urn:Rockwell/MasterRecipe}Parameter")
#         nm_el = etree.SubElement(new_el, "{urn:Rockwell/MasterRecipe}Name")
#         nm_el.text = row_dict.get("Name", "")

#         # Ensure ERPAlias is always created
#         erpalias_el = etree.SubElement(new_el, "{urn:Rockwell/MasterRecipe}ERPAlias")
#         erpalias_el.text = row_dict.get("ERPAlias", "").strip() or None  # Ensures empty tag if not present

#         # Track if a numerical tag exists
#         has_numeric_type = False
#         low_el = None  # Placeholder for <Low> element

#         # Fill sub-elements (only if non-blank)
#         created_sub_count = 0
#         for k, v in row_dict.items():
#             if k in ("FullPath", "TagType", "Name", "ERPAlias"):  # ERPAlias is already handled separately
#                 continue
#             if k.startswith("FormulaValueLimit_"):
#                 continue
#             if v.strip() == "":  # Skip blank values
#                 continue
            
#             sub_el = etree.SubElement(new_el, f"{{urn:Rockwell/MasterRecipe}}{k}")
#             sub_el.text = v
#             created_sub_count += 1

#             # Track <Integer>, <String>, or <Real>
#             if k in ("Integer", "String", "Real"):
#                 has_numeric_type = True
            
#             # Track the <Low> element to insert <EngineeringUnits> after it
#             if k == "Low":
#                 low_el = sub_el

#         # Ensure <EngineeringUnits/> after <Low> if an Integer, String, or Real exists
#         if has_numeric_type and low_el is not None:
#             eng_units_el = etree.Element("{urn:Rockwell/MasterRecipe}EngineeringUnits")
#             new_el.insert(new_el.index(low_el) + 1, eng_units_el)

#         new_parameters.append(new_el)
#         stats["created_count"] += 1
#         logger.debug(f"Created Parameter with {created_sub_count} sub-elements: {fp}")

#     elif last_part.startswith("FormulaValue["):
#         # Locate step for FormulaValue
#         step_el = locate_step_in_dom(root, middle)
#         if step_el is None:
#             logger.warning(f"Cannot locate step for {fp}, skipping create.")
#             return

#         # Create new formula value element
#         new_el = etree.Element("{urn:Rockwell/MasterRecipe}FormulaValue")
#         nm_el = etree.SubElement(new_el, "{urn:Rockwell/MasterRecipe}Name")
#         nm_el.text = row_dict.get("Name", "")
#         apply_formulavalue_updates(new_el, row_dict, original_info=None)
#         new_formula_values.append(new_el)
#         stats["created_count"] += 1
#         logger.debug(f"Created FormulaValue: {fp}")
#     else:
#         logger.warning(f"Unrecognized path => {fp}")

#     # Reorder elements in the tree
#     # Find the last Parameter element in the tree
#     last_param_el = None
#     for param in root.iter("{urn:Rockwell/MasterRecipe}Parameter"):
#         last_param_el = param  # Keep track of the last Parameter

#     if new_parameters:
#         if last_param_el is not None:
#             # Insert the new parameter right after the last existing one
#             root.insert(root.index(last_param_el) + 1, new_parameters[0])
#         else:
#             # If no parameter exists, just append the new parameter
#             root.append(new_parameters[0])

#     # Append all formula values after parameters
#     for formula in new_formula_values:
#         root.append(formula)


def create_parameter_or_formulavalue(row_dict, filepath, tree, node_info, stats):
    fp = row_dict.get("FullPath", "")
    logger.info(f"Creating node in {filepath}: {fp}")

    root = tree.getroot()
    parts = fp.split("/")
    last_part = parts[-1]
    middle = parts[1:-1]  # Used for locating parent step for FormulaValue

    if last_part.startswith("Parameter["):
        # Create new Parameter element
        new_el = etree.Element("{urn:Rockwell/MasterRecipe}Parameter")

        # Name (required)
        name_el = etree.SubElement(new_el, "{urn:Rockwell/MasterRecipe}Name")
        name_el.text = row_dict.get("Name", "")

        # ERPAlias (always created, even if blank)
        erpalias_el = etree.SubElement(new_el, "{urn:Rockwell/MasterRecipe}ERPAlias")
        erpalias_el.text = row_dict.get("ERPAlias", "").strip() or None

        # Flags
        has_numeric_type = False
        low_el = None
        created_sub_count = 0

        for k, v in row_dict.items():
            if k in ("FullPath", "TagType", "Name", "ERPAlias"):
                continue
            if k.startswith("FormulaValueLimit_") or not v.strip():
                continue

            sub_el = etree.SubElement(new_el, f"{{urn:Rockwell/MasterRecipe}}{k}")
            sub_el.text = v
            created_sub_count += 1

            if k in ("Integer", "String", "Real"):
                has_numeric_type = True
            if k == "Low":
                low_el = sub_el

        # if has_numeric_type and low_el is not None:
        #     eng_units_el = etree.Element("{urn:Rockwell/MasterRecipe}EngineeringUnits")
        #     new_el.insert(new_el.index(low_el) + 1, eng_units_el)

        if has_numeric_type:
            eng_units_el = etree.Element("{urn:Rockwell/MasterRecipe}EngineeringUnits")
            if low_el is not None:
                new_el.insert(new_el.index(low_el) + 1, eng_units_el)
            else:
                new_el.append(eng_units_el)

        last_param_el = None
        for param in root.iter("{urn:Rockwell/MasterRecipe}Parameter"):
            last_param_el = param

        if last_param_el is not None:
            root.insert(root.index(last_param_el) + 1, new_el)
        else:
            root.append(new_el)

        stats["created_count"] += 1
        logger.debug(f"Created Parameter with {created_sub_count} sub-elements: {fp}")

    elif last_part.startswith("FormulaValue["):
        step_el = locate_step_in_dom(root, middle)
        if step_el is None:
            logger.warning(f"Cannot locate step for {fp}, skipping create.")
            return

        # Create FormulaValue element
        new_el = etree.Element("{urn:Rockwell/MasterRecipe}FormulaValue")

        # Preferred order of child tags
        preferred_order = [
            "Name",
            "Display",
            "Defer",
            "EnumerationSet",
            "EnumerationMember"
        ]

        # Create elements in preferred order
        for tag in preferred_order:
            value = row_dict.get(tag, "").strip()
            sub_el = etree.SubElement(new_el, f"{{urn:Rockwell/MasterRecipe}}{tag}")
            if value:
                sub_el.text = value

        # Add any additional fields not in preferred order
        for k, v in row_dict.items():
            if k in ("FullPath", "TagType") or k in preferred_order or not v.strip():
                continue
            sub_el = etree.SubElement(new_el, f"{{urn:Rockwell/MasterRecipe}}{k}")
            sub_el.text = v

        # Append to step
        step_el.append(new_el)

        stats["created_count"] += 1
        logger.debug(f"Created FormulaValue inside step: {fp}")

    else:
        logger.warning(f"Unrecognized path => {fp}")




def apply_formulavalue_updates(fv_element, row_dict, original_info):
    """
    If there's no original_info => creation from scratch
    If there is original_info => partial update
    """
    ns = fv_element.nsmap

    # handle normal sub elements first
    if original_info is not None:
        # partial update
        original_subs = original_info["OriginalSubs"]  # localName->text
        for k, new_val in row_dict.items():
            if k in ("FullPath", "TagType", "Name"):
                continue
            if k.startswith("FormulaValueLimit_"):
                continue
            old_val = original_subs.get(k, None)
            if new_val.strip() == "":
                # user blank
                if old_val is None:
                    # didn't exist => skip
                    continue
                else:
                    # existed => preserve => do nothing
                    pass
            else:
                # user gave something => update or create
                sub_el = fv_element.find(
                    f"{{urn:Rockwell/MasterRecipe}}{k}", namespaces=ns
                )
                if sub_el is None:
                    sub_el = etree.SubElement(
                        fv_element, f"{{urn:Rockwell/MasterRecipe}}{k}"
                    )
                sub_el.text = new_val
    else:
        # creation from scratch => only add if non blank
        for k, new_val in row_dict.items():
            if k in ("FullPath", "TagType", "Name"):
                continue
            if k.startswith("FormulaValueLimit_"):
                continue
            if new_val.strip() == "":
                continue
            sub_el = etree.SubElement(fv_element, f"{{urn:Rockwell/MasterRecipe}}{k}")
            sub_el.text = new_val

    # handle FormulaValueLimit (FVL)
    fvl_keys = [k for k in row_dict if k.startswith("FormulaValueLimit_")]
    if not fvl_keys:
        return

    if original_info is None:
        # creation from scratch => only create FVL if any value is non-blank
        any_non_blank = any(row_dict[k].strip() != "" for k in fvl_keys)
        if not any_non_blank:
            return
        fvl_el = etree.SubElement(
            fv_element, "{urn:Rockwell/MasterRecipe}FormulaValueLimit"
        )
        # fill it
        for k in fvl_keys:
            v = row_dict[k]
            suffix = k[len("FormulaValueLimit_") :]
            if suffix == "Verification":
                if v.strip():
                    fvl_el.set("Verification", v)
            else:
                if v.strip():
                    sub_sub = etree.SubElement(
                        fvl_el, f"{{urn:Rockwell/MasterRecipe}}{suffix}"
                    )
                    sub_sub.text = v
    else:
        # partial update
        had_fvl = original_info["HasFormulaValueLimit"]
        old_ver = original_info["FormulaValueLimitVerification"]
        old_subs = original_info["FormulaValueLimitSubs"]
        fvl_el = fv_element.find(
            "{urn:Rockwell/MasterRecipe}FormulaValueLimit", namespaces=ns
        )

        if not had_fvl and fvl_el is None:
            # only create if user has non-blank
            any_non_blank = any(row_dict[k].strip() != "" for k in fvl_keys)
            if any_non_blank:
                fvl_el = etree.SubElement(
                    fv_element, "{urn:Rockwell/MasterRecipe}FormulaValueLimit"
                )

        # if had_fvl => it should exist, else we preserve old
        if had_fvl and fvl_el is None:
            # recreate if somehow missing
            fvl_el = etree.SubElement(
                fv_element, "{urn:Rockwell/MasterRecipe}FormulaValueLimit"
            )
            if old_ver:
                fvl_el.set("Verification", old_ver)
            for subk, subv in old_subs.items():
                sub_sub = etree.SubElement(
                    fvl_el, f"{{urn:Rockwell/MasterRecipe}}{subk}"
                )
                sub_sub.text = subv

        if fvl_el is None:
            return  # do nothing if still absent

        # update keys
        for k in fvl_keys:
            v = row_dict[k]
            suffix = k[len("FormulaValueLimit_") :]
            if suffix == "Verification":
                if v.strip() == "":
                    # user blank => preserve if old
                    if had_fvl:
                        pass
                else:
                    fvl_el.set("Verification", v)
            else:
                old_v = old_subs.get(suffix, None) if had_fvl else None
                if v.strip() == "":
                    if old_v is None:
                        pass
                    else:
                        pass
                else:
                    sub_sub = fvl_el.find(
                        f"{{urn:Rockwell/MasterRecipe}}{suffix}", namespaces=ns
                    )
                    if sub_sub is None:
                        sub_sub = etree.SubElement(
                            fvl_el, f"{{urn:Rockwell/MasterRecipe}}{suffix}"
                        )
                    sub_sub.text = v


def update_parameter_or_formulavalue(node_info, row_dict, filepath, stats):
    el = node_info["Element"]
    tagtype = node_info["TagType"]
    if tagtype == "Parameter":
        update_parameter(el, row_dict, node_info)
        logger.debug(f"Updated Parameter: {row_dict.get('FullPath', '')}")
    else:
        apply_formulavalue_updates(el, row_dict, original_info=node_info)
        logger.debug(f"Updated FormulaValue: {row_dict.get('FullPath', '')}")
    stats["updated_count"] += 1


def update_parameter(param_element, row_dict, node_info):
    ns = param_element.nsmap
    original_subs = node_info["OriginalSubs"]  # localName->text
    for k, new_val in row_dict.items():
        if k in ("FullPath", "TagType", "Name"):
            continue
        if k.startswith("FormulaValueLimit_"):
            continue
        old_val = original_subs.get(k, None)
        if new_val.strip() == "":
            # user blank
            if old_val is None:
                # didn't exist => skip
                continue
            else:
                # existed => preserve old => do nothing
                pass
        else:
            # user gave something => set or create
            sub_el = param_element.find(
                f"{{urn:Rockwell/MasterRecipe}}{k}", namespaces=ns
            )
            if sub_el is None:
                sub_el = etree.SubElement(
                    param_element, f"{{urn:Rockwell/MasterRecipe}}{k}"
                )
            sub_el.text = new_val


def command_excel2xml(parent_xml, excel_path):
    logger.info(f"Starting excel2xml: parent={parent_xml}, excel={excel_path}")
    if not os.path.isfile(parent_xml):
        logger.error(f"Parent file not found: {parent_xml}")
        sys.exit(1)
    if not os.path.isfile(excel_path):
        logger.error(f"Excel file not found: {excel_path}")
        sys.exit(1)

    file_to_tree = load_original_trees(parent_xml)
    # build fullpath maps
    file_to_fullpath_map = {}
    for fpath, tree in file_to_tree.items():
        fm = {}
        build_fullpath_map_for_file(tree, fpath, fm)
        file_to_fullpath_map[fpath] = fm

    # read excel
    wb = load_workbook(excel_path)
    sheet_to_data = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = rows[0]
        data_rows = rows[1:]
        this_sheet_data = {}
        for dr in data_rows:
            row_dict = {}
            for col_idx, col_name in enumerate(headers):
                if not col_name:
                    continue
                val = dr[col_idx] if col_idx < len(dr) else ""
                if val is None:
                    val = ""
                row_dict[col_name] = str(val)
            fp = row_dict.get("FullPath", "")
            if fp:
                this_sheet_data[fp] = row_dict
        sheet_to_data[sheet] = this_sheet_data

    stats = {"deleted_count": 0, "created_count": 0, "updated_count": 0}

    # apply changes
    for fpath, tree in file_to_tree.items():
        short_name = os.path.splitext(os.path.basename(fpath))[0]
        excel_data_for_file = sheet_to_data.get(short_name, {})
        existing_map = file_to_fullpath_map[fpath]

        excel_fullpaths = set(excel_data_for_file.keys())
        existing_fullpaths = set(existing_map.keys())

        to_delete = existing_fullpaths - excel_fullpaths
        to_create = excel_fullpaths - existing_fullpaths
        to_update = existing_fullpaths.intersection(excel_fullpaths)

        # remove
        for fp in to_delete:
            remove_parameter_or_formulavalue(existing_map[fp], fpath, stats)

        # create
        for fp in to_create:
            row_d = excel_data_for_file[fp]
            create_parameter_or_formulavalue(
                row_d, fpath, tree, node_info=None, stats=stats
            )

        # update
        for fp in to_update:
            ni = existing_map[fp]
            row_d = excel_data_for_file[fp]
            update_parameter_or_formulavalue(ni, row_d, fpath, stats)

    # final logs
    logger.info(
        f"Deleted: {stats['deleted_count']}, Created: {stats['created_count']}, Updated: {stats['updated_count']}"
    )

    # write out
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d-%H%M%S")
    outdir = os.path.join(os.path.dirname(parent_xml), "converted-outputs", timestamp)
    os.makedirs(outdir, exist_ok=True)

    for fpath, tree in file_to_tree.items():
        outpath = os.path.join(outdir, os.path.basename(fpath))
        logger.info(f"Writing updated XML => {outpath}")
        tree.write(outpath, encoding="UTF-8", xml_declaration=True, pretty_print=False)

    logger.info("excel2xml complete. All updated XML saved.")


def main():
    parser = argparse.ArgumentParser(description="FactoryTalk Batch Bulk Editor")

    sub = parser.add_subparsers(dest="command", help="Sub-command")

    p1 = sub.add_parser(
        "xml2excel", help="Export .pxml/.uxml/.oxml + children to Excel"
    )
    p1.add_argument("--xml", required=True, help="XML .pxml/.uxml/.oxml file")
    p1.add_argument("--excel", required=True, help="Output .xlsx path")
    p1.add_argument(
        "--debug",
        action="store_true",  # on/off flag
        required=False,
        default=False,
        help="Enable debug logging to file (batch_bulk_editor.log)",
    )
    p2 = sub.add_parser(
        "excel2xml", help="Import Excel changes back to .pxml/.uxml/.oxml"
    )
    p2.add_argument("--xml", required=True, help="XML .pxml/.uxml/.oxml")
    p2.add_argument("--excel", required=True, help="Edited .xlsx path")
    p2.add_argument(
        "--debug",
        action="store_true",  # on/off flag
        required=False,
        default=False,
        help="Enable debug logging to file (batch_bulk_editor.log)",
    )
    args = parser.parse_args()

    debug_mode = getattr(args, "debug", False)
    configure_logging(debug_mode)

    if args.command == "xml2excel":
        command_xml2excel(args.xml, args.excel)
    elif args.command == "excel2xml":
        command_excel2xml(args.xml, args.excel)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
