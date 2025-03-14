# utils/excel_converter.py
from openpyxl import Workbook, load_workbook
from models.recipe import (
    RecipeElement, Header, Parameter, Steps, Step, FormulaValue,
    PhaseLinkGroup, PhaseLink, Transition, ElementLink,
    UnitRequirement, DownstreamResource, Formulations, Formulation, FormulationParameter
)
from typing import List

class ExcelConverter:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file

    def recipe_to_excel(self, recipe: RecipeElement):
        wb = Workbook()

        # --- Header Sheet ---
        ws_header = wb.active
        ws_header.title = "Header"
        ws_header.append(["Field", "Value"])
        for field, value in recipe.header.dict().items():
            ws_header.append([field, value])

        # --- Parameters Sheet ---
        ws_params = wb.create_sheet("Parameters")
        param_headers = list(Parameter.__fields__.keys())
        ws_params.append(param_headers)
        for param in recipe.parameters:
            row = [getattr(param, field) for field in param_headers]
            ws_params.append(row)

        # --- Steps Sheet ---
        ws_steps = wb.create_sheet("Steps")
        # We'll add a column "type" to indicate whether the step is initial, terminal, or regular.
        step_headers = ["type", "name", "x_pos", "y_pos", "acquire_unit", "system_step", "step_recipe_id", "packed_flags", "unit_alias"]
        ws_steps.append(step_headers)
        if recipe.steps:
            if recipe.steps.initial_step:
                s = recipe.steps.initial_step
                ws_steps.append(["initial_step", s.name, s.x_pos, s.y_pos, s.acquire_unit, s.system_step, s.step_recipe_id, s.packed_flags, s.unit_alias])
            if recipe.steps.terminal_step:
                s = recipe.steps.terminal_step
                ws_steps.append(["terminal_step", s.name, s.x_pos, s.y_pos, s.acquire_unit, s.system_step, s.step_recipe_id, s.packed_flags, s.unit_alias])
            for s in recipe.steps.steps:
                ws_steps.append(["step", s.name, s.x_pos, s.y_pos, s.acquire_unit, s.system_step, s.step_recipe_id, s.packed_flags, s.unit_alias])

        # --- PhaseLinkGroup Sheet ---
        ws_plg = wb.create_sheet("PhaseLinkGroup")
        ws_plg.append(["name"])
        if recipe.phase_link_group:
            ws_plg.append([recipe.phase_link_group.name])
            # --- PhaseLinks Sheet ---
            ws_pl = wb.create_sheet("PhaseLinks")
            pl_headers = list(PhaseLink.__fields__.keys())
            ws_pl.append(pl_headers)
            for pl in recipe.phase_link_group.phase_links:
                ws_pl.append([getattr(pl, field) for field in pl_headers])

        # --- Transitions Sheet ---
        ws_trans = wb.create_sheet("Transitions")
        trans_headers = list(Transition.__fields__.keys())
        ws_trans.append(trans_headers)
        for trans in recipe.transitions:
            ws_trans.append([getattr(trans, field) for field in trans_headers])

        # --- ElementLinks Sheet ---
        ws_elinks = wb.create_sheet("ElementLinks")
        elink_headers = list(ElementLink.__fields__.keys())
        ws_elinks.append(elink_headers)
        for elink in recipe.element_links:
            ws_elinks.append([getattr(elink, field) for field in elink_headers])

        # --- UnitRequirements Sheet ---
        ws_units = wb.create_sheet("UnitRequirements")
        # For UnitRequirement we include a column for downstream_resource name separately.
        unit_headers = ["unit_alias", "class_instance", "binding_method", "material_binding_method", "class_based", "downstream_resource_name"]
        ws_units.append(unit_headers)
        for ur in recipe.unit_requirements:
            dr_name = ur.downstream_resource.name if ur.downstream_resource else ""
            ws_units.append([ur.unit_alias, ur.class_instance, ur.binding_method, ur.material_binding_method, ur.class_based, dr_name])

        # --- Comments Sheet ---
        ws_comments = wb.create_sheet("Comments")
        ws_comments.append(["comments"])
        ws_comments.append([recipe.comments])

        # --- Formulations Sheet ---
        ws_forms = wb.create_sheet("Formulations")
        form_headers = ["name", "description"]
        ws_forms.append(form_headers)
        if recipe.formulations:
            for form in recipe.formulations.formulations:
                ws_forms.append([form.name, form.description])
                # Create a dedicated sheet for each formulation's parameters (limit sheet name length)
                sheet_name = f"Form_{form.name[:25]}"
                ws_form_params = wb.create_sheet(sheet_name)
                fp_headers = list(FormulationParameter.__fields__.keys())
                ws_form_params.append(fp_headers)
                for fp in form.parameter_list:
                    ws_form_params.append([getattr(fp, field) for field in fp_headers])

        wb.save(self.excel_file)
        print(f"Excel file saved to {self.excel_file}")

    def excel_to_recipe(self, original_recipe: RecipeElement) -> RecipeElement:
        wb = load_workbook(self.excel_file)

        # --- Header ---
        if "Header" in wb.sheetnames:
            ws_header = wb["Header"]
            header_dict = {}
            for row in ws_header.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    header_dict[row[0]] = row[1]
            original_recipe.header = Header(**header_dict)

        # --- Parameters ---
        if "Parameters" in wb.sheetnames:
            ws_params = wb["Parameters"]
            headers = [cell for cell in next(ws_params.iter_rows(min_row=1, max_row=1, values_only=True))]
            parameters = []
            for row in ws_params.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                lowercase_list = [s.lower() for s in headers]
                row_dict = dict(zip(lowercase_list, row))#headers, row))
                parameters.append(Parameter(**row_dict))
            original_recipe.parameters = parameters

        # --- Steps ---
        if "Steps" in wb.sheetnames:
            ws_steps = wb["Steps"]
            steps_obj = Steps()
            for row in ws_steps.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                step_type, name, x_pos, y_pos, acquire_unit, system_step, step_recipe_id, packed_flags, unit_alias = row
                step_obj = Step(
                    name=name,
                    x_pos=int(x_pos) if x_pos is not None else None,
                    y_pos=int(y_pos) if y_pos is not None else None,
                    acquire_unit=(str(acquire_unit).lower() == "true") if acquire_unit is not None else None,
                    system_step=(str(system_step).lower() == "true") if system_step is not None else None,
                    step_recipe_id=step_recipe_id,
                    packed_flags=int(packed_flags) if packed_flags is not None else None,
                    unit_alias=unit_alias,
                    formula_values=[]
                )
                if step_type == "initial_step":
                    steps_obj.initial_step = step_obj
                elif step_type == "terminal_step":
                    steps_obj.terminal_step = step_obj
                else:
                    steps_obj.steps.append(step_obj)
            original_recipe.steps = steps_obj

        # --- PhaseLinkGroup and PhaseLinks ---
        if "PhaseLinkGroup" in wb.sheetnames:
            ws_plg = wb["PhaseLinkGroup"]
            rows = list(ws_plg.iter_rows(min_row=2, values_only=True))
            if rows:
                plg_name = rows[0][0]
                phase_link_group = PhaseLinkGroup(name=plg_name, phase_links=[])
                if "PhaseLinks" in wb.sheetnames:
                    ws_pl = wb["PhaseLinks"]
                    pl_headers = [cell for cell in next(ws_pl.iter_rows(min_row=1, max_row=1, values_only=True))]
                    phase_links = []
                    for row in ws_pl.iter_rows(min_row=2, values_only=True):
                        if not any(row):
                            continue
                        row_dict = dict(zip(pl_headers, row))
                        phase_links.append(PhaseLink(**row_dict))
                    phase_link_group.phase_links = phase_links
                original_recipe.phase_link_group = phase_link_group

        # --- Transitions ---
        if "Transitions" in wb.sheetnames:
            ws_trans = wb["Transitions"]
            trans_headers = [cell for cell in next(ws_trans.iter_rows(min_row=1, max_row=1, values_only=True))]
            transitions = []
            for row in ws_trans.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = dict(zip(trans_headers, row))
                transitions.append(Transition(**row_dict))
            original_recipe.transitions = transitions

        # --- ElementLinks ---
        if "ElementLinks" in wb.sheetnames:
            ws_elinks = wb["ElementLinks"]
            elink_headers = [cell for cell in next(ws_elinks.iter_rows(min_row=1, max_row=1, values_only=True))]
            element_links = []
            for row in ws_elinks.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = dict(zip(elink_headers, row))
                element_links.append(ElementLink(**row_dict))
            original_recipe.element_links = element_links

        # --- UnitRequirements ---
        if "UnitRequirements" in wb.sheetnames:
            ws_units = wb["UnitRequirements"]
            unit_headers = [cell for cell in next(ws_units.iter_rows(min_row=1, max_row=1, values_only=True))]
            unit_requirements = []
            for row in ws_units.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = dict(zip(unit_headers, row))
                dr = None
                if "downstream_resource_name" in row_dict and row_dict["downstream_resource_name"]:
                    dr = DownstreamResource(name=row_dict["downstream_resource_name"])
                unit_requirements.append(UnitRequirement(
                    unit_alias=row_dict.get("unit_alias"),
                    class_instance=row_dict.get("class_instance"),
                    binding_method=row_dict.get("binding_method"),
                    material_binding_method=row_dict.get("material_binding_method"),
                    class_based=row_dict.get("class_based"),
                    downstream_resource=dr
                ))
            original_recipe.unit_requirements = unit_requirements

        # --- Comments ---
        if "Comments" in wb.sheetnames:
            ws_comments = wb["Comments"]
            rows = list(ws_comments.iter_rows(min_row=2, values_only=True))
            if rows and rows[0]:
                original_recipe.comments = rows[0][0]

        # --- Formulations ---
        if "Formulations" in wb.sheetnames:
            ws_forms = wb["Formulations"]
            form_headers = [cell for cell in next(ws_forms.iter_rows(min_row=1, max_row=1, values_only=True))]
            formulations_list = []
            for row in ws_forms.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = dict(zip(form_headers, row))
                form_name = row_dict.get("name")
                description = row_dict.get("description")
                param_list = []
                sheet_name = f"Form_{form_name[:25]}"
                if sheet_name in wb.sheetnames:
                    ws_form_params = wb[sheet_name]
                    fp_headers = [cell for cell in next(ws_form_params.iter_rows(min_row=1, max_row=1, values_only=True))]
                    for fp_row in ws_form_params.iter_rows(min_row=2, values_only=True):
                        if not any(fp_row):
                            continue
                        fp_dict = dict(zip(fp_headers, fp_row))
                        param_list.append(FormulationParameter(**fp_dict))
                formulations_list.append(Formulation(name=form_name, description=description, parameter_list=param_list))
            original_recipe.formulations = Formulations(formulations=formulations_list)

        return original_recipe
