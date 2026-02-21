"""
ExcelImporter: apply changes from an Excel workbook into RecipeTree models.
"""

import os
import logging
from collections.abc import Callable
from collections import defaultdict
from openpyxl import load_workbook
from utils.string import safe_strip
from utils.errors import ValidationError


class ExcelImporter:
    """
    Import changes from Excel workbook into RecipeTree instances.
    """

    def import_changes(
        self,
        excel_path: str,
        trees: list,
        progress_cb: Callable[[str, dict], None] | None = None,
    ):
        """
        Read an edited Excel workbook and apply create/update/delete operations to the RecipeTree models.

        This method opens the given `.xlsx` file and matches each sheet by filename to its
        corresponding RecipeTree.  It then iterates every data row, enforcing that exactly one
        data-type column is populated and that any `Defer` references exist in the same shee's
        parameter names.  Rows for existing nodes invoke `update_from_dict()`, while new rows
        invoke `create_parameter()` or `create_formulavalue()`.  After processing all rows, it
        removes any XML nodes that were not seen in Excel, counting them as deleted.  If any
        validation errors occur (e.g. missing defer target or type conflicts), it logs each as
        ERROR and raises a single `ValidationError` summarizing them.  A dict of counts
        (`created`, `updated`, `deleted`).

        Args:
            excel_path: Path to the edited Excel workbook.
            trees: List of RecipeTree loaded via XMLParser.parse().

        Raises:
            ValidationError: Aggregated validation errors found in Excel rows.
        """
        log = logging.getLogger(__name__)
        debug_enabled = log.isEnabledFor(logging.DEBUG)  # noqa: F841

        def _emit(event: str, **payload) -> None:
            if progress_cb is not None:
                progress_cb(event, payload)

        wb = load_workbook(excel_path)
        errors = []

        # Overall stats of total changes to xmls
        stats = {
            "created": 0,
            "updated": 0,
            "deleted": 0,
        }

        # detailed per‐sheet
        detailed = {}

        log.debug("Importing Excel changes to XML(s)")

        # Map sheet name to RecipeTree by basename
        tree_map = {os.path.basename(t.filepath): t for t in trees}
        total_sheets = len(wb.sheetnames)
        _emit("start", total=total_sheets)

        for index, sheet in enumerate(wb.sheetnames, start=1):
            if sheet not in tree_map:
                log.warning("No matching XML for sheet '%s', skipping", sheet)
                _emit("sheet_skipped", index=index, total=total_sheets, sheet=sheet)
                continue

            log.debug(f"Sheet {sheet}")
            _emit("sheet_start", index=index, total=total_sheets, sheet=sheet)

            # initialize sheet‐level stats
            sheet_stats = {
                "Parameters": {"Created": 0, "Updated": 0, "Deleted": 0},
                "FormulaValues": defaultdict(
                    lambda: {"Created": 0, "Updated": 0, "Deleted": 0, "Deferrals": 0}
                ),
            }
            detailed[sheet] = sheet_stats

            tree = tree_map[sheet]
            ws = wb[sheet]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

            excel_param_names = set()
            seen_params = set()
            seen_fvs = set()

            for r_idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                row_dict = dict(zip(header, [v if v is not None else "" for v in row]))
                fp = row_dict.get("FullPath", "").strip()
                tagtype = row_dict.get("TagType", "").strip()

                # Validate single-type
                types = ["Real", "Integer", "String", "EnumerationSet", "Defer"]
                cnt = sum(bool(safe_strip(row_dict.get(t))) for t in types)
                if cnt > 2:
                    log.error(
                        f"{sheet}!Row{r_idx}: expected exactly one data type for {fp}"
                    )
                    errors.append(
                        f"{sheet}!Row{r_idx}: expected exactly one data type for {fp}"
                    )
                    continue

                if tagtype == "Parameter":
                    excel_param_names.add(row_dict["Name"].strip())
                    node = tree.find_parameter(fp)
                    log.debug(f"\tWorking on Parameter: {row_dict["Name"].strip()}")
                    if node:
                        if node.update_from_dict(row_dict):
                            log.debug(
                                "\tParameter found in XML (updating)"  # updating data to: {row_dict=}"
                            )
                            stats["updated"] += 1
                            sheet_stats["Parameters"]["Updated"] += 1
                        else:
                            log.debug(
                                "\tParameter found in XML (no changes found)"  # updating data to: {row_dict=}"
                            )
                    else:
                        tree.create_parameter(row_dict)
                        log.debug(
                            "\tParameter NOT found in XML (creating)"  # creating parameter with: {row_dict=}"
                        )
                        stats["created"] += 1
                        sheet_stats["Parameters"]["Created"] += 1
                    seen_params.add(fp)

                elif tagtype == "FormulaValue":
                    node = tree.find_formulavalue(fp)
                    defer = row_dict.get("Defer", "").strip()
                    log.debug(f"\tWorking on FormulaValue: {row_dict["Name"].strip()}")

                    # determine step for stats
                    step = fp.split("/Steps/Step[", 1)[1].split("]")[0]

                    # if defer and not tree.has_parameter_named(defer):
                    if defer and defer not in excel_param_names:
                        log.error(
                            f"{sheet}!Row{r_idx}: defer target '{defer}' not found for {fp}"
                        )
                        errors.append(
                            f"{sheet}!Row{r_idx}: defer target '{defer}' not found for {fp}"
                        )
                        continue
                    if node:
                        if node.update_from_dict(row_dict):
                            log.debug(
                                "\tFormulaValue found in XML (updating)"  # , updating data to: {row_dict=}"
                            )
                            stats["updated"] += 1
                            sheet_stats["FormulaValues"][step]["Updated"] += 1
                            if defer:
                                sheet_stats["FormulaValues"][step]["Deferrals"] += 1
                        else:
                            log.debug(
                                "\tFormulaValue found in XML (no change)"  # , updating data to: {row_dict=}"
                            )
                    else:
                        tree.create_formulavalue(row_dict)
                        log.debug(
                            "\tFormulaValue NOT found in XML (creating)"  # , creating with: {row_dict=}"
                        )
                        stats["created"] += 1
                        sheet_stats["FormulaValues"][step]["Created"] += 1
                        if defer:
                            sheet_stats["FormulaValues"][step]["Deferrals"] += 1
                    seen_fvs.add(fp)
                else:
                    log.error(f"{sheet}!Row{r_idx}: unknown TagType '{tagtype}'")
                    errors.append(f"{sheet}!Row{r_idx}: unknown TagType '{tagtype}'")
                    continue

            # Deletes
            for node in list(tree.parameters):
                if node.fullpath not in seen_params:
                    node.element.getparent().remove(node.element)
                    tree.parameters.remove(node)
                    log.debug(
                        f"\tParameter NOT found in Excel but exists in XML, {node.fullpath} deleted.."
                    )
                    stats["deleted"] += 1
                    sheet_stats["Parameters"]["Deleted"] += 1
            for node in list(tree.formula_values):
                if node.fullpath not in seen_fvs:
                    node.element.getparent().remove(node.element)
                    tree.formula_values.remove(node)
                    step = node.fullpath.split("/Steps/Step[", 1)[1].split("]")[0]

                    log.debug(
                        f"\tFormulaValue NOT found in Excel but exists in XML, {node.fullpath} deleted.."
                    )
                    stats["deleted"] += 1
                    sheet_stats["FormulaValues"][step]["Deleted"] += 1

            # per‐sheet summary
            p = sheet_stats["Parameters"]
            fv_stats = sheet_stats["FormulaValues"]
            any_sheet_changes = (
                p["Created"]
                or p["Updated"]
                or p["Deleted"]
                or any(
                    f["Created"] or f["Updated"] or f["Deleted"]
                    for f in fv_stats.values()
                )
            )

            if any_sheet_changes:
                log.debug("Changes for '%s'", sheet)
                log.debug(
                    "Parameters → Created=%d\tUpdated=%d\tDeleted=%d",
                    p["Created"],
                    p["Updated"],
                    p["Deleted"],
                )
                log.debug("FormulaValues by step:")
                for step, f in sheet_stats["FormulaValues"].items():
                    log.debug(
                        "\t[%s] Created=%d\tUpdated=%d\tOut of which %d are Deferrals\tDeleted=%d",
                        step,
                        f["Created"],
                        f["Updated"],
                        f["Deferrals"],
                        f["Deleted"],
                    )
                log.debug("------------------------------------------------")
            sheet_created = p["Created"] + sum(
                f["Created"] for f in sheet_stats["FormulaValues"].values()
            )
            sheet_updated = p["Updated"] + sum(
                f["Updated"] for f in sheet_stats["FormulaValues"].values()
            )
            sheet_deleted = p["Deleted"] + sum(
                f["Deleted"] for f in sheet_stats["FormulaValues"].values()
            )
            _emit(
                "sheet_done",
                index=index,
                total=total_sheets,
                sheet=sheet,
                created=sheet_created,
                updated=sheet_updated,
                deleted=sheet_deleted,
            )

        if errors:
            raise ValidationError(f"{len(errors)} errors")

        # overall summary
        log.debug(
            "Workbook import summary: created=%d updated=%d deleted=%d",
            stats["created"],
            stats["updated"],
            stats["deleted"],
        )
        log.debug("-- Summary of changes made by excel2xml Tool ---")
        log.debug("------------------------------------------------")
        log.debug(
            "Total → Created=%d\tUpdated=%d\tDeleted=%d",
            stats["created"],
            stats["updated"],
            stats["deleted"],
        )
        log.debug("------------------------------------------------")
        _emit("finished", total=total_sheets, stats=stats)
        return stats
